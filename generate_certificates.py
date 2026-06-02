import openpyxl
import zipfile
import os
import re
import img2pdf
from PIL import Image, ImageDraw, ImageFont
import arabic_reshaper
from bidi.algorithm import get_display

EXCEL    = "/root/.claude/uploads/d38d69a7-e04c-423f-a3b7-6b9589550bcd/52ee1c80-______.xlsx"
TEMPLATE = "/root/.claude/uploads/08d794ce-260f-42e6-8f06-48c7dfe21b84/7fbfe9b9-tempalte.docx"
OUT_DIR  = "/home/user/Test/certificates"

# Page in twips: 16838 x 11906 (landscape A4)
# EMU per twip = 914400 / 1440 = 635
PAGE_W_EMU = 16838 * 635   # 10,692,130
PAGE_H_EMU = 11906 * 635   # 7,559,810
IMG_W, IMG_H = 2229, 1551  # background image pixels

# Arabic font
FONT_AR = "/usr/share/fonts/truetype/noto/NotoNaskhArabic-Bold.ttf"
FONT_EN = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
FONT_EN_BOLD = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"

CITY_EN = {
    'الرياض': 'Riyadh', 'الدمام': 'Dammam', 'جدة': 'Jeddah',
    'الجبيل': 'Jubail', 'المدينة': 'Madinah', 'أبها': 'Abha',
    'تبوك': 'Tabuk', 'القصيم': 'Qassim', 'جازان': 'Jazan',
    'حائل': 'Hail', 'الأحساء': 'Al-Ahsa', 'ينبع': 'Yanbu',
}

def emu_to_px(emu_x, emu_y):
    return int(emu_x / PAGE_W_EMU * IMG_W), int(emu_y / PAGE_H_EMU * IMG_H)

def ar(text):
    """Reshape and apply bidi to Arabic text for correct rendering."""
    return get_display(arabic_reshaper.reshape(text))

def draw_centered(draw, text, cx, cy, font, fill):
    bbox = draw.textbbox((0, 0), text, font=font)
    w = bbox[2] - bbox[0]
    draw.text((cx - w // 2, cy), text, font=font, fill=fill)

def feminize_arabic_role(role):
    if role.startswith('منسق ') and not role.startswith('منسقة'):
        return 'منسقة ' + role[len('منسق '):]
    if role == 'منسق':
        return 'منسقة'
    return role

def replace_in_xml(xml, placeholder, value):
    return xml.replace(f'<w:t>{placeholder}</w:t>', f'<w:t>{value}</w:t>')

def apply_gender_xml(xml, male):
    if male:
        xml = xml.replace('>للدكتورة<', '>للدكتور<')
        xml = xml.replace('>جهودها<', '>جهوده<')
        xml = xml.replace('>في عملها<', '>في عمله<')
        xml = xml.replace('>her outstanding efforts<', '>his outstanding efforts<')
        xml = xml.replace('>her role as <', '>his role as <')
    return xml

def apply_center_xml(xml, center, male):
    section_ar = 'الطلاب' if male else 'الطالبات'
    section_en = 'Male Section' if male else 'Female Section'
    city_en = CITY_EN.get(center.strip(), center.strip())
    if center:
        ar_center = f' بالمركز التعليمي ب{center} لشطر {section_ar} خلال العام الجامعي ١٤٤٧ هـ'
        en_center = f' at the {city_en} Educational Center, {section_en}, during the academic year 1447 AH.'
    else:
        ar_center = ' خلال العام الجامعي ١٤٤٧ هـ'
        en_center = ' during the academic year 1447 AH.'
    xml = xml.replace(
        '> بالمركز التعليمي بجدة لشطر الطالبات خلال العام الجامعي ١٤٤٧ هـ<',
        f'>{ar_center}<')
    xml = xml.replace(
        '> at the Jeddah Educational Center, Female Section, during the academic year 1447 AH.<',
        f'>{en_center}<')
    return xml

def generate_docx(arabic_name, english_name, arabic_role, english_role, male, center, out_path):
    with zipfile.ZipFile(TEMPLATE, 'r') as z:
        names = z.namelist()
        files = {name: z.read(name) for name in names}
    xml = files['word/document.xml'].decode('utf-8')
    xml = apply_gender_xml(xml, male)
    xml = apply_center_xml(xml, center, male)
    # Connect ك with role: replace standalone " ك" node before X
    xml = xml.replace('<w:t xml:space="preserve"> ك</w:t>', '<w:t></w:t>')
    xml = replace_in_xml(xml, 'A', arabic_name)
    xml = replace_in_xml(xml, 'X', 'ك' + arabic_role)
    xml = replace_in_xml(xml, 'C', english_name)
    xml = replace_in_xml(xml, 'Y', english_role)
    files['word/document.xml'] = xml.encode('utf-8')
    with zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name in names:
            zout.writestr(name, files[name])

def generate_pdf(docx_path, pdf_path, arabic_name, english_name,
                 arabic_role, english_role, male, center):
    # Load background image from the generated docx
    with zipfile.ZipFile(docx_path) as z:
        bg_data = z.read('word/media/image2.jpg')
        logo_data = z.read('word/media/image1.jpeg')

    import io
    bg = Image.open(io.BytesIO(bg_data)).convert('RGB')
    logo = Image.open(io.BytesIO(logo_data)).convert('RGBA')
    W, H = bg.size  # 2229 x 1551

    draw = ImageDraw.Draw(bg)

    # Fonts
    f_ar_big  = ImageFont.truetype(FONT_AR, 38)
    f_ar_med  = ImageFont.truetype(FONT_AR, 30)
    f_ar_sm   = ImageFont.truetype(FONT_AR, 24)
    f_en_big  = ImageFont.truetype(FONT_EN_BOLD, 32)
    f_en_med  = ImageFont.truetype(FONT_EN, 26)
    f_en_sm   = ImageFont.truetype(FONT_EN, 20)

    BLUE  = (0, 32, 96)
    TEAL  = (21, 96, 130)

    # ── Arabic section (center of page) ──────────────────────────────
    # Anchor 2: pos=(4556234, 2098347), size=(4744720, 3168650)
    ax, ay = emu_to_px(4556234 + 4744720//2, 2098347)  # center x of box
    # Lines
    draw_centered(draw, ar('يسر كلية الحوسبة والمعلوماتية بالجامعة السعودية الالكترونية'), ax, ay+10,  f_ar_sm, BLUE)
    draw_centered(draw, ar('أن تتقدم بوافر الشكر والتقدير'),                                ax, ay+55,  f_ar_sm, BLUE)

    # للدكتور/للدكتورة / Name
    title_ar = 'للدكتور' if male else 'للدكتورة'
    name_line_ar = ar(f'{title_ar} / {arabic_name}')
    draw_centered(draw, name_line_ar, ax, ay+105, f_ar_big, TEAL)

    # على جهوده/ا المميزة في عمله/ا كRole
    effort_ar = 'على جهوده المميزة في عمله' if male else 'على جهودها المميزة في عملها'
    draw_centered(draw, ar(effort_ar), ax, ay+165, f_ar_sm, BLUE)
    draw_centered(draw, ar('ك' + arabic_role), ax, ay+210, f_ar_med, TEAL)

    # Center phrase Arabic
    section_ar = 'الطلاب' if male else 'الطالبات'
    if center:
        draw_centered(draw, ar(f'بالمركز التعليمي ب{center} لشطر {section_ar}'), ax, ay+260, f_ar_sm, BLUE)
        draw_centered(draw, ar('خلال العام الجامعي ١٤٤٧ هـ'), ax, ay+305, f_ar_sm, BLUE)
    else:
        draw_centered(draw, ar('خلال العام الجامعي ١٤٤٧ هـ'), ax, ay+270, f_ar_sm, BLUE)

    draw_centered(draw, ar('سائلين المولى لها مزِيدًا من النجاح والتقدم' if not male
                           else 'سائلين المولى له مزِيدًا من النجاح والتقدم'),
                  ax, ay + (355 if center else 320), f_ar_sm, BLUE)

    # ── English section ───────────────────────────────────────────────
    # Anchor 3: pos=(~10187174-4824248/2, 2098106) — right half
    ex = int((PAGE_W_EMU - 4824248//2) / PAGE_W_EMU * W)
    ey = emu_to_px(0, 2098106)[1]

    section_en = 'Male Section' if male else 'Female Section'
    city_en = CITY_EN.get(center.strip(), center.strip()) if center else ''

    draw_centered(draw, 'The College of Computing and Informatics is pleased to', ex, ey+10, f_en_sm, BLUE)
    draw_centered(draw, 'extend its sincere thanks and appreciation to',           ex, ey+38, f_en_sm, BLUE)
    his_her = 'his' if male else 'her'
    draw_centered(draw, f'Dr. / {english_name}', ex, ey+85, f_en_big, TEAL)
    draw_centered(draw, f'for {his_her} outstanding efforts in {his_her} role as', ex, ey+140, f_en_sm, BLUE)
    draw_centered(draw, english_role, ex, ey+170, f_en_med, TEAL)
    if center:
        draw_centered(draw, f'at the {city_en} Educational Center, {section_en},', ex, ey+215, f_en_sm, BLUE)
        draw_centered(draw, 'during the academic year 1447 AH.',                    ex, ey+243, f_en_sm, BLUE)
    else:
        draw_centered(draw, 'during the academic year 1447 AH.',                    ex, ey+220, f_en_sm, BLUE)
    draw_centered(draw, 'Wishing continued success and excellence.', ex, ey+285, f_en_sm, BLUE)

    # Save to PDF via img2pdf
    img_bytes = io.BytesIO()
    bg.save(img_bytes, format='JPEG', quality=95)
    img_bytes.seek(0)
    with open(pdf_path, 'wb') as f:
        f.write(img2pdf.convert(img_bytes.read()))

def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active

    for row in range(7, ws.max_row + 1):
        arabic_name  = ws.cell(row, 11).value
        english_name = ws.cell(row, 10).value
        arabic_role  = ws.cell(row, 9).value
        english_role = ws.cell(row, 5).value
        gender       = ws.cell(row, 3).value
        center       = ws.cell(row, 1).value

        if not arabic_name and not english_name:
            continue

        arabic_name  = (arabic_name  or '').strip()
        english_name = (english_name or '').strip()
        arabic_role  = (arabic_role  or '').strip()
        english_role = (english_role or '').strip()
        center       = (center       or '').strip()
        male = str(gender or '').strip() == 'ذكر'

        if not male:
            arabic_role = feminize_arabic_role(arabic_role)

        safe = arabic_name.replace(' ', '_').replace('.', '').replace('/', '')
        docx_out = os.path.join(OUT_DIR, f'{safe}.docx')
        pdf_out  = os.path.join(OUT_DIR, f'{safe}.pdf')

        print(f"Generating: {arabic_name} ({'M' if male else 'F'}) center={center or 'none'}")
        generate_docx(arabic_name, english_name, arabic_role, english_role, male, center, docx_out)
        generate_pdf(docx_out, pdf_out, arabic_name, english_name,
                     arabic_role, english_role, male, center)

    print(f'\nDone! {OUT_DIR}')

if __name__ == '__main__':
    main()
