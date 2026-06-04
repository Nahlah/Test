import openpyxl
import zipfile
import os
import re
import subprocess

EXCEL     = "/root/.claude/uploads/85bb3a05-ff2f-4a4d-b7e8-9152bc5d2234/156bfaab-______.xlsx"
TEMPLATE_F = "/root/.claude/uploads/85bb3a05-ff2f-4a4d-b7e8-9152bc5d2234/a7fce754-tempalte_1_otherdocx.docx"
TEMPLATE_M = "/root/.claude/uploads/85bb3a05-ff2f-4a4d-b7e8-9152bc5d2234/863346f2-tempalte_2other.docx"
OUT_DIR   = "/home/user/Test/certificates"

ROLE_TRANSLATIONS = {
    'منسقة الأنشطة الطلابية':                      'Coordinator of Student Activities',
    'منسق الأنشطة الطلابية':                       'Coordinator of Student Activities',
    'لجنة شؤون الطلاب/ الأنشطة الطلابية':          'Student Affairs Committee / Student Activities',
    'لجنة شؤون الطلاب/الأنشطة الطلابية':           'Student Affairs Committee / Student Activities',
    'منسق فرع':                                    'Branch Coordinator',
    'منسقة فرع':                                   'Branch Coordinator',
    'منسق برنامج':                                  'Program Coordinator',
    'منسقة برنامج':                                 'Program Coordinator',
    'منسق الاختبارات':                              'Exam Coordinator',
    'منسقة الاختبارات':                             'Exam Coordinator',
    'منسقة مقررات السنة المشتركة':                  'Common Year Courses Coordinator',
    'منسقة مشاريع التخرج لبرامج البكالوريوس':       'Graduation Projects Coordinator for Bachelor Programs',
    'مشرفة برنامج ماجستير علوم البيانات':           "Supervisor of the Master's Program in Data Science",
    'مشرف برنامج ماجستير الأمن السيبراني':          "Supervisor of the Master's Program in Cybersecurity",
    'منسقة الأنشطة الطلابية':                      'Coordinator of Student Activities',
    'منسق الأنشطة الطلابية':                       'Coordinator of Student Activities',
    'لجنة الأنشطة الطلابية':                       'Student Activities Committee',

    'منسقة الإرشاد الأكاديمي':                     'Academic Advising Coordinator',
    'منسق الإرشاد الأكاديمي':                      'Academic Advising Coordinator',
    'لجنة الإرشاد والإشراف الدراسي':               'Academic Advising and Supervision Committee',
    'لجنة التدريب التعاوني':                        'Cooperative Training Committee',
    'أعمال إدارية':                                 'Administrative Work',
    'منسقة خدمة المجتمع':                          'Community Service Coordinator',
    'منسق خدمة المجتمع':                           'Community Service Coordinator',
    'أمينة لجنة الانضباط الفرعية':                 'Secretary of the Branch Disciplinary Committee',
    'أمين لجنة الانضباط الفرعية':                  'Secretary of the Branch Disciplinary Committee',
    'بلجنة الأنشطة الطلابية':                      'Student Activities Committee',
    'بلجنة الأنشطة الطلابية':                      'Student Activities Committee',
    'بلجنة الإرشاد والإشراف الدراسي':              'Academic Advising and Supervision Committee',
    'بلجنة التدريب التعاوني':                       'Cooperative Training Committee',
    'أعمال إدارية':                                 'Administrative Work',
}

def translate_role(arabic_role):
    return ROLE_TRANSLATIONS.get(arabic_role.strip(), arabic_role.strip())

def replace_in_xml(xml, placeholder, value):
    return xml.replace(f'<w:t>{placeholder}</w:t>', f'<w:t>{value}</w:t>')

def generate_docx(template, arabic_name, arabic_title, english_name,
                  english_prefix, arabic_role, out_path):
    english_role = translate_role(arabic_role)

    with zipfile.ZipFile(template, 'r') as z:
        names = z.namelist()
        files = {n: z.read(n) for n in names}

    xml = files['word/document.xml'].decode('utf-8')

    xml = replace_in_xml(xml, 'A', arabic_name)
    xml = replace_in_xml(xml, 'B', arabic_title)
    xml = replace_in_xml(xml, 'C', english_name)
    xml = replace_in_xml(xml, 'D', english_prefix)
    xml = replace_in_xml(xml, 'X', arabic_role)
    xml = replace_in_xml(xml, 'Y', english_role)

    # Remove bottom padding from text boxes ending with "Wishing" and "سائلين"
    for phrase in ['Wishing', 'سائلين']:
        idx = xml.find(phrase)
        if idx != -1:
            para_end = xml.find('</w:p>', idx) + 6
            bodypr_start = xml.find('<wps:bodyPr', para_end)
            bodypr_end = xml.find('/>', bodypr_start) + 2
            old_bodypr = xml[bodypr_start:bodypr_end]
            new_bodypr = old_bodypr.replace('bIns="45720"', 'bIns="0"')
            xml = xml[:bodypr_start] + new_bodypr + xml[bodypr_end:]

    files['word/document.xml'] = xml.encode('utf-8')

    with zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            zout.writestr(n, files[n])

def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active

    generated = 0
    skipped   = 0

    for row in range(7, ws.max_row + 1):
        gender       = ws.cell(row, 2).value   # col B
        arabic_role  = ws.cell(row, 7).value   # col G = X
        english_name = ws.cell(row, 8).value   # col H = C
        arabic_name  = ws.cell(row, 9).value   # col I = A
        eng_prefix   = ws.cell(row, 10).value  # col J = D
        arabic_title = ws.cell(row, 11).value  # col K = B
        num          = ws.cell(row, 12).value  # col L = #

        # Skip rows missing essential data
        if not arabic_name or not english_name:
            if num:
                print(f"  ⚠ Row {row} (#{num}): missing name — skipped")
            skipped += 1
            continue

        arabic_name  = arabic_name.strip()
        english_name = english_name.strip()
        arabic_role  = ' '.join((arabic_role  or '').split())
        arabic_title = (arabic_title or '').strip()
        eng_prefix   = (eng_prefix   or '').strip()
        male = str(gender or '').strip().startswith('ذكر')

        template = TEMPLATE_M if male else TEMPLATE_F

        safe = arabic_name.replace(' ', '_').replace('.', '').replace('/', '')
        safe_role = arabic_role.replace(' ', '_').replace('/', '')
        out_path = os.path.join(OUT_DIR, f'{safe}_{safe_role}.docx')

        print(f"Generating: {arabic_name} ({'M' if male else 'F'})")
        generate_docx(template, arabic_name, arabic_title, english_name,
                      eng_prefix, arabic_role, out_path)
        generated += 1

    print(f"\nDone: {generated} generated, {skipped} skipped (incomplete data).")

if __name__ == '__main__':
    main()
