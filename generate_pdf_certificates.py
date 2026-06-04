import openpyxl
import fitz
import arabic_reshaper
from bidi.algorithm import get_display
import os

EXCEL      = "/root/.claude/uploads/74144cc4-c35c-4349-9618-cb72d129c09c/ca39c260-______.xlsx"
TEMPLATE_F = "/root/.claude/uploads/74144cc4-c35c-4349-9618-cb72d129c09c/75a9dfab-tempalte_1_otherdocx.pdf"
TEMPLATE_M = "/root/.claude/uploads/74144cc4-c35c-4349-9618-cb72d129c09c/63f8520a-tempalte_2other.pdf"
OUT_DIR    = "/home/user/Test/certificates"

ROLE_TRANSLATIONS = {
    'منسقة الأنشطة الطلابية':               'Student Activities Committee',
    'منسق الأنشطة الطلابية':                'Student Activities Committee',
    'لجنة الأنشطة الطلابية':                'Student Activities Committee',
    'لجنة  الأنشطة الطلابية':               'Student Activities Committee',
    'بلجنة الأنشطة الطلابية':               'Student Activities Committee',
    'لجنة الإرشاد والإشراف الدراسي':        'Academic Advising and Supervision Committee',
    'بلجنة الإرشاد والإشراف الدراسي':       'Academic Advising and Supervision Committee',
    'منسقة الإرشاد الأكاديمي':              'Academic Advising Coordinator',
    'منسق الإرشاد الأكاديمي':               'Academic Advising Coordinator',
    'لجنة التدريب التعاوني':                 'Cooperative Training Committee',
    'بلجنة التدريب التعاوني':                'Cooperative Training Committee',
    'أعمال إدارية':                          'Administrative Work',
    'منسقة خدمة المجتمع':                   'Community Service Coordinator',
    'منسق خدمة المجتمع':                    'Community Service Coordinator',
    'أمينة لجنة الانضباط الفرعية':          'Secretary of the Branch Disciplinary Committee',
    'أمين لجنة الانضباط الفرعية':           'Secretary of the Branch Disciplinary Committee',
    'لجنة شؤون الطلاب/ الأنشطة الطلابية':  'Student Affairs Committee / Student Activities',
    'لجنة شؤون الطلاب/الأنشطة الطلابية':   'Student Affairs Committee / Student Activities',
    'منسق فرع':                             'Branch Coordinator',
    'منسقة فرع':                            'Branch Coordinator',
    'منسق برنامج':                           'Program Coordinator',
    'منسقة برنامج':                          'Program Coordinator',
    'منسق الاختبارات':                       'Exam Coordinator',
    'منسقة الاختبارات':                      'Exam Coordinator',
}

def translate_role(role):
    return ROLE_TRANSLATIONS.get(role.strip(), role.strip())

def ar(text):
    return get_display(arabic_reshaper.reshape(text))

def cover_and_write(page, rect, text, fontsize, color, align_right=False):
    page.draw_rect(rect, color=(1, 1, 1), fill=(1, 1, 1))
    if align_right:
        # insert from right edge
        page.insert_text(
            (rect.x1 - 2, rect.y1 + fontsize * 0.85),
            text, fontsize=fontsize, color=color,
            fontname="helv"
        )
    else:
        page.insert_text(
            (rect.x0, rect.y1 - 2),
            text, fontsize=fontsize, color=color,
            fontname="helv"
        )

def generate_pdf(template_path, arabic_name, arabic_title, english_name,
                 english_prefix, arabic_role, out_path, male):
    english_role = translate_role(arabic_role)
    doc = fitz.open(template_path)
    page = doc[0]

    BLUE = (0, 0.12, 0.38)
    TEAL = (0.08, 0.38, 0.51)
    FS = 13.9  # font size matching template

    # ── English side ────────────────────────────────────────────────────
    # Replace "D. / C" with actual prefix / name
    dc_rect = fitz.Rect(47, 256, 350, 285) if not male else fitz.Rect(58, 263, 420, 292)
    en_line = f"{english_prefix}. / {english_name}"
    cover_and_write(page, dc_rect, en_line, FS, TEAL)

    # Replace the "for her/his ... Y" line
    if male:
        effort_rect = fitz.Rect(58, 290, 700, 318)
        effort_text = f"for his outstanding efforts in his role as a member of the {english_role}"
    else:
        effort_rect = fitz.Rect(47, 283, 700, 311)
        effort_text = f"for her outstanding efforts in her role as a member of the {english_role}"
    cover_and_write(page, effort_rect, effort_text, FS, BLUE)

    # ── Arabic side ─────────────────────────────────────────────────────
    # Replace "A /B" with arabic_name / arabic_title
    ab_rect = fitz.Rect(440, 256, 800, 285) if not male else fitz.Rect(440, 263, 800, 292)
    ab_text = ar(f'{arabic_name} / {arabic_title}')
    cover_and_write(page, ab_rect, ab_text, FS, TEAL, align_right=True)

    # Replace "X" with arabic_role
    x_rect = fitz.Rect(200, 282, 490, 311) if not male else fitz.Rect(200, 290, 497, 318)
    x_text = ar(arabic_role)
    cover_and_write(page, x_rect, x_text, FS, TEAL, align_right=True)

    doc.save(out_path)
    doc.close()

def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb.active

    generated = skipped = 0

    for row in range(7, ws.max_row + 1):
        gender       = ws.cell(row, 2).value
        arabic_role  = ws.cell(row, 7).value
        english_name = ws.cell(row, 8).value
        arabic_name  = ws.cell(row, 9).value
        eng_prefix   = ws.cell(row, 10).value
        arabic_title = ws.cell(row, 11).value
        num          = ws.cell(row, 12).value

        if not arabic_name or not english_name:
            if num:
                print(f"  ⚠ Row {row} (#{num}): missing name — skipped")
            skipped += 1
            continue

        arabic_name  = arabic_name.strip()
        english_name = english_name.strip()
        arabic_role  = ' '.join((arabic_role  or '').split())
        arabic_title = (arabic_title or '').strip()
        eng_prefix   = (eng_prefix   or '').strip()
        male = str(gender or '').strip().startswith('ذكر')

        template = TEMPLATE_M if male else TEMPLATE_F
        safe = arabic_name.replace(' ', '_').replace('.', '').replace('/', '')
        out_path = os.path.join(OUT_DIR, f'{safe}.pdf')

        print(f"Generating: {arabic_name} ({'M' if male else 'F'})")
        generate_pdf(template, arabic_name, arabic_title, english_name,
                     eng_prefix, arabic_role, out_path, male)
        generated += 1

    print(f"\nDone: {generated} generated, {skipped} skipped.")

if __name__ == '__main__':
    main()
